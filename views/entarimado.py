import streamlit as st
import pandas as pd
import datetime
from utils.database import get_connection, save_db_to_excel, sync_and_push_db

def get_inventario_pt():
    conn = get_connection()
    # 1. Obtener avances acumulados en EMPAQUE y en LIBERADO por pieza
    #    Se usa el MAYOR de los dos para cubrir piezas que pasaron calidad
    #    pero no tienen registro explícito de Empaque (ruta sin Empaque o avance faltante).
    df_avances = pd.read_sql_query("""
        SELECT 
            a.of_number,
            a.no_pieza,
            p.descripcion,
            COALESCE(SUM(CASE WHEN a.area = 'Empaque' THEN a.cantidad ELSE 0 END), 0)  as total_empaque,
            COALESCE(SUM(CASE WHEN a.area = 'Liberado' THEN a.cantidad ELSE 0 END), 0) as total_liberado
        FROM avances a
        LEFT JOIN (
            SELECT of_number, no_pieza, MIN(nombre_pieza) as descripcion
            FROM piezas
            GROUP BY of_number, no_pieza
        ) p ON a.of_number = p.of_number AND a.no_pieza = p.no_pieza
        WHERE a.area IN ('Empaque', 'Liberado')
        GROUP BY a.of_number, a.no_pieza
    """, conn)
    
    # 2. Obtener lo ya entarimado
    df_tarimas = pd.read_sql_query("""
        SELECT of_number, no_pieza, SUM(cantidad) as total_empaquetado
        FROM tarimas
        GROUP BY of_number, no_pieza
    """, conn)
    
    # 3. Obtener metadata de ordenes (PO, Proyecto, etc.)
    df_meta = pd.read_sql_query("""
        SELECT of_number, po, proyecto, proyecto_cliente, descripcion_pronest
        FROM ordenes
    """, conn)
    conn.close()
    
    if df_avances.empty:
        return pd.DataFrame(columns=[
            "OF", "Producto/SKU", "Descripción", "PO", "Proyecto", 
            "Proyecto Cliente", "Total Avanzado en PT", "Total Entarimado", "Disponible en PT",
            "Fuente PT"
        ])
    
    # Usar el mayor de (Empaque, Liberado) como referencia de piezas en PT
    df_avances["total_avances"] = df_avances[["total_empaque", "total_liberado"]].max(axis=1).astype(int)
    df_avances["Fuente PT"] = df_avances.apply(
        lambda r: "Empaque" if r["total_empaque"] >= r["total_liberado"] else "Liberado", axis=1
    )
        
    df_inv = df_avances.merge(df_tarimas, on=["of_number", "no_pieza"], how="left")
    df_inv["total_empaquetado"] = df_inv["total_empaquetado"].fillna(0).astype(int)
    df_inv["disponible"] = df_inv["total_avances"] - df_inv["total_empaquetado"]
    
    df_inv = df_inv.merge(df_meta, on="of_number", how="left")
    
    df_inv.rename(columns={
        "of_number": "OF",
        "no_pieza": "Producto/SKU",
        "descripcion": "Descripción",
        "po": "PO",
        "proyecto": "Proyecto",
        "proyecto_cliente": "Proyecto Cliente",
        "total_avances": "Total Avanzado en PT",
        "total_empaquetado": "Total Entarimado",
        "disponible": "Disponible en PT"
    }, inplace=True)
    
    df_inv = df_inv[[
        "OF", "Producto/SKU", "Descripción", "PO", "Proyecto", "Proyecto Cliente", 
        "Total Avanzado en PT", "Total Entarimado", "Disponible en PT", "Fuente PT"
    ]]
    return df_inv


def get_next_bulto_name():
    conn = get_connection()
    c = conn.cursor()
    c.execute("SELECT DISTINCT tarima_id FROM tarimas")
    rows = c.fetchall()
    conn.close()
    
    max_num = 0
    for r in rows:
        name = r[0]
        if name.startswith("Bulto_"):
            try:
                num = int(name.split("_")[1])
                if num > max_num:
                    max_num = num
            except ValueError:
                pass
    return f"Bulto_{max_num + 1}"

def generate_plantilla_tarimas_excel(selected_tarima_ids):
    import openpyxl
    import io
    import os
    
    base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    template_path = os.path.join(base_dir, "plantilla_carga_tarimas_sigrama.xlsx")
    wb = openpyxl.load_workbook(template_path)
    ws = wb['Plantilla_Tarimas']
    
    # Limpiar datos anteriores (desde fila 2 hasta la última con datos)
    if ws.max_row > 1:
        for r in range(2, ws.max_row + 1):
            for c in range(1, 8):
                ws.cell(row=r, column=c).value = None
                
    conn = get_connection()
    placeholders = ",".join(["?"] * len(selected_tarima_ids))
    cursor = conn.cursor()
    cursor.execute(f"""
        SELECT t.tarima_id, t.no_pieza, o.po, 
               COALESCE(o.proyecto_cliente, o.proyecto) as proyecto, 
               o.prioridad, 
               o.descripcion_pronest, 
               t.cantidad
        FROM tarimas t
        LEFT JOIN ordenes o ON t.of_number = o.of_number
        WHERE t.tarima_id IN ({placeholders})
    """, selected_tarima_ids)
    rows = cursor.fetchall()
    conn.close()
    
    from openpyxl.styles import Alignment
    from openpyxl.utils import get_column_letter
    center_align = Alignment(horizontal="center", vertical="center")
    
    for i, row in enumerate(rows):
        row_num = i + 2
        ws.cell(row=row_num, column=1, value=row[0]) # Tarima
        ws.cell(row=row_num, column=2, value=row[1]) # Producto/SKU
        ws.cell(row=row_num, column=3, value=row[2] if row[2] else "") # PO
        ws.cell(row=row_num, column=4, value=row[3] if row[3] else "") # Proyecto
        ws.cell(row=row_num, column=5, value=row[4] if row[4] else "") # Parcialidad
        ws.cell(row=row_num, column=6, value=row[5] if row[5] else "") # Descripcion
        ws.cell(row=row_num, column=7, value=row[6]) # Cantidad
        
        # Aplicar alineación centrada a las celdas de datos
        for col_num in range(1, 8):
            ws.cell(row=row_num, column=col_num).alignment = center_align
            
    # Auto-ajustar el ancho de las columnas basándose en el contenido
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        # Dar un margen extra de 5 caracteres, con un mínimo de 12
        ws.column_dimensions[col_letter].width = max(max_len + 5, 12)
        
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()

def get_historial_movimientos():
    """Obtiene el historial de movimientos de entarimado ordenado por fecha."""
    conn = get_connection()
    df = pd.read_sql_query("""
        SELECT 
            t.timestamp as "Fecha",
            t.tarima_id as "Lote",
            t.of_number as "OF",
            t.no_pieza as "Producto/SKU",
            COALESCE(p.nombre_pieza, '') as "Descripción",
            COALESCE(o.po, '') as "PO",
            COALESCE(o.proyecto_cliente, o.proyecto, '') as "Proyecto",
            t.cantidad as "Cantidad Entarimada"
        FROM tarimas t
        LEFT JOIN (
            SELECT of_number, no_pieza, MIN(nombre_pieza) as nombre_pieza
            FROM piezas GROUP BY of_number, no_pieza
        ) p ON t.of_number = p.of_number AND t.no_pieza = p.no_pieza
        LEFT JOIN ordenes o ON t.of_number = o.of_number
        ORDER BY t.timestamp DESC
        LIMIT 200
    """, conn)
    conn.close()
    return df


def generate_historial_excel(df):
    """Genera Excel del historial de movimientos."""
    import io, openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Historial Entarimado"

    # Header style
    hdr_fill = PatternFill("solid", fgColor="1A1A2E")
    hdr_font = Font(color="FFFFFF", bold=True, size=11)
    center = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin", color="DDDDDD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = list(df.columns)
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = center
        cell.border = border

    # Data rows
    alt_fill = PatternFill("solid", fgColor="F5F5F5")
    for row_idx, row in enumerate(df.itertuples(index=False), 2):
        fill = alt_fill if row_idx % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
        for col_idx, val in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.fill = fill
            cell.alignment = center
            cell.border = border

    # Auto column widths
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max(max_len + 4, 12)

    ws.row_dimensions[1].height = 22
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def view_entarimado():
    st.markdown("## 📦 CONTROL DE ENTARIMADO (PRODUCTO TERMINADO)")

    # ── Sección 1: Inventario editable ──────────────────────────────────────
    st.markdown("### 1. Inventario Disponible en Almacén PT")
    st.caption("Marca las piezas que ya están entarimadas y ajusta la cantidad si es necesario.")

    df_inv = get_inventario_pt()
    df_disp = df_inv[df_inv["Disponible en PT"] > 0].copy().reset_index(drop=True)

    if df_disp.empty:
        st.info("⚠️ No hay piezas disponibles en Almacén PT. Primero registra avances en 'Liberado' o 'Empaque'.")
    else:
        # Preparar tabla editable: agregar columnas de acción
        df_edit = df_disp[[
            "OF", "Producto/SKU", "Descripción", "PO", "Proyecto",
            "Total Avanzado en PT", "Total Entarimado", "Disponible en PT", "Fuente PT"
        ]].copy()
        df_edit.insert(0, "✅ Entarimar", False)
        df_edit.insert(1, "Cantidad", df_edit["Disponible en PT"].astype(int))

        edited = st.data_editor(
            df_edit,
            use_container_width=True,
            hide_index=True,
            key="entarimado_editor",
            column_config={
                "✅ Entarimar": st.column_config.CheckboxColumn(
                    "✅ Entarimar",
                    help="Marca para registrar estas piezas como entarimadas",
                    default=False,
                    width="small"
                ),
                "Cantidad": st.column_config.NumberColumn(
                    "Cantidad",
                    help="Piezas a registrar como entarimadas",
                    min_value=1,
                    step=1,
                    width="small"
                ),
                "OF": st.column_config.TextColumn("OF", width="medium"),
                "Producto/SKU": st.column_config.TextColumn("Producto/SKU", width="small"),
                "Descripción": st.column_config.TextColumn("Descripción", width="large"),
                "PO": st.column_config.TextColumn("PO", width="small"),
                "Proyecto": st.column_config.TextColumn("Proyecto", width="small"),
                "Total Avanzado en PT": st.column_config.NumberColumn("Total PT", width="small"),
                "Total Entarimado": st.column_config.NumberColumn("Entarimado", width="small"),
                "Disponible en PT": st.column_config.NumberColumn("Disponible", width="small"),
                "Fuente PT": st.column_config.TextColumn("Fuente", width="small"),
            },
            disabled=[
                "OF", "Producto/SKU", "Descripción", "PO", "Proyecto",
                "Total Avanzado en PT", "Total Entarimado", "Disponible en PT", "Fuente PT"
            ],
            num_rows="fixed",
        )

        # Filas marcadas
        seleccionadas = edited[edited["✅ Entarimar"] == True]

        if not seleccionadas.empty:
            st.markdown(f"**{len(seleccionadas)} pieza(s) seleccionada(s) para entarimar:**")

            # Validación en tiempo real
            errores = []
            for _, row in seleccionadas.iterrows():
                cant = int(row["Cantidad"])
                disp = int(row["Disponible en PT"])
                if cant <= 0:
                    errores.append(f"❌ `{row['Producto/SKU']}`: la cantidad debe ser mayor a 0.")
                elif cant > disp:
                    errores.append(f"❌ `{row['Producto/SKU']}`: cantidad ({cant}) supera el disponible ({disp}).")

            for e in errores:
                st.error(e)

            if not errores:
                if st.button("📦 Confirmar Entarimado", type="primary", use_container_width=True):
                    conn = get_connection()
                    cursor = conn.cursor()
                    now_str = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    lote_id = f"ENT-{datetime.datetime.now().strftime('%Y%m%d-%H%M%S')}"

                    for _, row in seleccionadas.iterrows():
                        cursor.execute("""
                            INSERT INTO tarimas (tarima_id, no_pieza, of_number, cantidad, timestamp)
                            VALUES (?, ?, ?, ?, ?)
                        """, (lote_id, row["Producto/SKU"], row["OF"], int(row["Cantidad"]), now_str))

                    conn.commit()
                    conn.close()

                    save_db_to_excel()
                    sync_and_push_db()

                    total_pzas = seleccionadas["Cantidad"].sum()
                    st.success(f"✅ ¡{int(total_pzas)} piezas registradas como entarimadas! Lote: `{lote_id}`")
                    st.rerun()
        else:
            st.info("☝️ Selecciona una o más piezas marcando la casilla **✅ Entarimar** y ajusta la cantidad si es necesario.")

    # ── Sección 2: Historial de movimientos ─────────────────────────────────
    st.markdown("---")
    st.markdown("### 2. Últimos Movimientos de Entarimado")

    df_hist = get_historial_movimientos()

    if df_hist.empty:
        st.info("No hay movimientos registrados aún.")
    else:
        col_info, col_dl = st.columns([3, 1])
        with col_info:
            st.caption(f"Mostrando los últimos {len(df_hist)} movimientos registrados.")
        with col_dl:
            excel_bytes = generate_historial_excel(df_hist)
            st.download_button(
                label="📥 Descargar Excel",
                data=excel_bytes,
                file_name=f"historial_entarimado_{datetime.datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                type="primary"
            )

        # Mostrar historial
        st.dataframe(
            df_hist,
            use_container_width=True,
            hide_index=True,
            height=400,
            column_config={
                "Fecha": st.column_config.TextColumn("Fecha", width="medium"),
                "Lote": st.column_config.TextColumn("Lote / ID", width="medium"),
                "OF": st.column_config.TextColumn("OF", width="medium"),
                "Producto/SKU": st.column_config.TextColumn("Producto/SKU", width="small"),
                "Descripción": st.column_config.TextColumn("Descripción", width="large"),
                "PO": st.column_config.TextColumn("PO", width="small"),
                "Proyecto": st.column_config.TextColumn("Proyecto", width="small"),
                "Cantidad Entarimada": st.column_config.NumberColumn("Piezas", width="small"),
            }
        )

        # Resumen rápido por OF
        st.markdown("#### Resumen por OF")
        resumen = df_hist.groupby(["OF", "Producto/SKU"])["Cantidad Entarimada"].sum().reset_index()
        resumen.columns = ["OF", "Producto/SKU", "Total Entarimado"]
        resumen = resumen.sort_values("OF")
        st.dataframe(resumen, use_container_width=True, hide_index=True, height=200)

