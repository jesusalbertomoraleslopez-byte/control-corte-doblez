"""
Script para generar la Plantilla Excel oficial de SIGRAMA
Formato nuevo: 3 hojas — Orden, Nidos, Piezas
"""
import io
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

# ─────────────────── ESTILOS ───────────────────
YELLOW_FILL = PatternFill("solid", fgColor="FFD700")
ORANGE_FILL = PatternFill("solid", fgColor="FFA500")
GRAY_FILL   = PatternFill("solid", fgColor="D2D3D5")
RED_FILL    = PatternFill("solid", fgColor="EC2024")
DARK_FILL   = PatternFill("solid", fgColor="111111")
WHITE_FILL  = PatternFill("solid", fgColor="FFFFFF")
LIGHT_FILL  = PatternFill("solid", fgColor="F5F5F5")
INFO_FILL   = PatternFill("solid", fgColor="E8F4FD")

def bold_font(size=11, color="000000", italic=False):
    return Font(name="Calibri", bold=True, size=size, color=color, italic=italic)

def normal_font(size=10, color="000000"):
    return Font(name="Calibri", bold=False, size=size, color=color)

def thin_border():
    s = Side(style="thin", color="AAAAAA")
    return Border(left=s, right=s, top=s, bottom=s)

def thick_border():
    s = Side(style="medium", color="111111")
    return Border(left=s, right=s, top=s, bottom=s)

def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def left_align():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

def apply_header(cell, text, fill, font_color="000000", size=11):
    cell.value = text
    cell.fill = fill
    cell.font = bold_font(size=size, color=font_color)
    cell.alignment = center()
    cell.border = thick_border()


# ─────────────────── HOJA 1: ORDEN ───────────────────
def build_orden_sheet(ws):
    ws.title = "Orden"
    ws.sheet_view.showGridLines = False

    # Anchos
    for col, w in [("A",22),("B",20),("C",22),("D",22),("E",16),("F",25),("G",14),("H",14),("I",25)]:
        ws.column_dimensions[col].width = w

    ws.row_dimensions[1].height = 38
    ws.row_dimensions[2].height = 24

    # Fila 1: Encabezados
    headers = [
        "Nombre del Proyecto *",
        "Fecha de Producción",
        "Programador *",
        "Orden de Fabricación",
        "PO",
        "Descripción de OF Pronest",
        "Calibre",
        "PRIORIDAD",
        "Nombre del Proyecto de Cliente",
    ]
    fills = [YELLOW_FILL] * len(headers)
    for col_idx, (h, f) in enumerate(zip(headers, fills), start=1):
        cell = ws.cell(row=1, column=col_idx)
        apply_header(cell, h, f, font_color="111111", size=11)

    # Fila 2: Datos de ejemplo
    ejemplos = ["PRUEBA 1", "2026/06/30", "BRYAN FLORES", "OF-00001", "2602-0711", "CORTE Y DOBLEZ DE PIEZAS", "Cal 14", 4, "CLIENTE SIGRAMA"]
    for c_idx, val in enumerate(ejemplos, start=1):
        cell = ws.cell(row=2, column=c_idx, value=val)
        cell.fill = PatternFill("solid", fgColor="FFFACD")
        cell.border = thin_border()
        cell.alignment = center()
        cell.font = normal_font(size=10, color="888888")

    # Filas vacías para uso futuro (3–10)
    for row in range(3, 11):
        for col in range(1, 10):
            cell = ws.cell(row=row, column=col)
            cell.fill = WHITE_FILL
            cell.border = thin_border()
            cell.alignment = center()
            cell.font = normal_font()


# ─────────────────── HOJA 2: NIDOS ───────────────────
def build_nidos_sheet(ws):
    ws.title = "Nidos"
    ws.sheet_view.showGridLines = False

    for col, w in [("A",14),("B",12)]:
        ws.column_dimensions[col].width = w

    ws.row_dimensions[1].height = 28

    # Fila 1: Encabezados
    headers = ["NIDO", "HOJAS"]
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        apply_header(cell, h, YELLOW_FILL, font_color="111111", size=11)

    # Filas de ejemplo (2–4) con N01, N02, N03
    ejemplos = [
        ["N01", 6],
        ["N02", 1],
        ["N03", 20],
    ]
    for r_idx, row_data in enumerate(ejemplos, start=2):
        fill = PatternFill("solid", fgColor="FFFACD")
        for c_idx, val in enumerate(row_data, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.fill  = fill
            cell.border = thin_border()
            cell.alignment = center()
            cell.font  = normal_font(size=10, color="888888")

    # Filas de datos vacías (5–55)
    for row in range(5, 56):
        fill = LIGHT_FILL if row % 2 == 0 else WHITE_FILL
        for col in range(1, 3):
            cell = ws.cell(row=row, column=col)
            cell.fill = fill
            cell.border = thin_border()
            cell.alignment = center()
            cell.font = normal_font()

    # Fila TOTAL (57)
    tc = ws.cell(row=57, column=1, value="TOTAL")
    tc.font  = bold_font(size=12, color="FFFFFF")
    tc.fill  = DARK_FILL
    tc.alignment = center()
    tc.border = thick_border()

    cell_nidos = ws.cell(row=57, column=2, value="=COUNTA(A2:A55)")
    cell_nidos.font  = bold_font(size=11)
    cell_nidos.fill  = GRAY_FILL
    cell_nidos.alignment = center()
    cell_nidos.border = thick_border()

    # Etiquetas bajo TOTAL (58)
    for col, label in [(1,"Nidos"),(2,"Hojas")]:
        lc = ws.cell(row=58, column=col, value=label)
        lc.font  = bold_font(size=9, color="FFFFFF")
        lc.fill  = DARK_FILL
        lc.alignment = center()
        lc.border = thin_border()


# ─────────────────── HOJA 3: PIEZAS ───────────────────
def build_piezas_sheet(ws):
    ws.title = "Piezas"
    ws.sheet_view.showGridLines = False

    for col, w in [("A",10),("B",18),("C",55),("D",12)]:
        ws.column_dimensions[col].width = w

    ws.row_dimensions[1].height = 28

    # Fila 1: Encabezados
    headers = ["NIDO", "No. PIEZA", "NOMBRE DE PIEZA", "CANTIDAD"]
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        apply_header(cell, h, YELLOW_FILL, font_color="111111", size=11)

    # Filas de ejemplo (2–6)
    ejemplos = [
        ["N01", "P19500-16", "P19500-16-(14gacr ANSI-61) - K48 - V3-R0", 5],
        ["N02", "P19500-16", "P19500-16-(14gacr ANSI-61) - K48 - V3-R0", 2],
        ["N02", "P17690-17", "P17690-17-(14gacr ANSI-61) - K48 - V1-R0", 2],
        ["N03", "PP9247",    "PP9247-(14gacr ANSI-61) - K48 - V1-R0",    10],
        ["N03", "P13704-091","P13704-091-(14gacr ANSI-61) - K48 - V5-R0", 4],
    ]
    for r_idx, row_data in enumerate(ejemplos, start=2):
        fill = PatternFill("solid", fgColor="FFFACD")
        for c_idx, val in enumerate(row_data, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.fill  = fill
            cell.border = thin_border()
            cell.alignment = center() if c_idx != 3 else left_align()
            cell.font  = normal_font(size=10, color="888888")

    # Datos vacíos (7–207)
    for row in range(7, 208):
        fill = LIGHT_FILL if row % 2 == 0 else WHITE_FILL
        for col in range(1, 5):
            cell = ws.cell(row=row, column=col)
            cell.fill = fill
            cell.border = thin_border()
            cell.font = normal_font()
            cell.alignment = center() if col != 3 else left_align()


# ─────────────────── FUNCIÓN PRINCIPAL ───────────────────
def generar_plantilla() -> bytes:
    wb = Workbook()
    del wb[wb.sheetnames[0]]

    ws_orden  = wb.create_sheet("Orden")
    ws_nidos  = wb.create_sheet("Nidos")
    ws_piezas = wb.create_sheet("Piezas")

    build_orden_sheet(ws_orden)
    build_nidos_sheet(ws_nidos)
    build_piezas_sheet(ws_piezas)

    wb.active = ws_orden

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


if __name__ == "__main__":
    data = generar_plantilla()
    with open("Plantilla_Plan_Produccion_SIGRAMA.xlsx", "wb") as f:
        f.write(data)
    print("✅ Plantilla generada exitosamente.")
